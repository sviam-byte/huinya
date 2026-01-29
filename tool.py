#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import importlib
import importlib.util
import logging
import os
import warnings
from io import BytesIO
from itertools import chain, combinations, permutations
from typing import List, Optional

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import networkx as nx
import nolds
import numpy as np
import pandas as pd
import scipy.signal as signal
from hurst import compute_Hc
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats
from scipy.fft import fft
from scipy.signal import coherence, find_peaks
from scipy.spatial import cKDTree
from scipy.special import digamma
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
from statsmodels.graphics.tsaplots import plot_acf
from statsmodels.tsa.stattools import adfuller, grangercausalitytests
from statsmodels.tsa.vector_ar.var_model import VAR
from tqdm import tqdm


def configure_warnings(quiet: bool = False) -> None:
    """Настраивает предупреждения без глобального подавления."""
    warnings.filterwarnings(
        "ignore",
        category=FutureWarning,
        module="statsmodels.tsa.stattools",
    )
    warnings.filterwarnings(
        "ignore",
        message="nperseg = 256 is greater than input length",
    )
    if quiet:
        warnings.filterwarnings("ignore")


base_path = os.path.dirname(os.path.abspath(__file__))
save_folder = os.path.join(base_path, "TimeSeriesAnalysis")
os.makedirs(save_folder, exist_ok=True)

# Параметр регуляризации для избежания вырождения
REG_ALPHA = 1e-5

##############################################
# Настройки по умолчанию
##############################################
DEFAULT_MAX_LAG = 5
DEFAULT_K_MI = 5
DEFAULT_BINS = 8
DEFAULT_OUTLIER_Z = 5
DEFAULT_REGULARIZATION = 1e-8
DEFAULT_EMBED_DIM = 3
DEFAULT_EMBED_TAU = 1

STABLE_METHODS = [
    "correlation_full",
    "correlation_partial",
    "coherence_full",
    "granger_full",
]

# Флаг наличия pyinform: нужен для расчёта transfer entropy.
PYINFORM_AVAILABLE = importlib.util.find_spec("pyinform") is not None

EXPERIMENTAL_METHODS_BASE = [
    "mutinf_full",
    "mutinf_partial",
    "te_full",
    "te_partial",
    "te_directed",
    "ah_full",
    "ah_partial",
    "ah_directed",
]

# Если pyinform недоступен, скрываем TE-методы из списка UI.
EXPERIMENTAL_METHODS = [
    method
    for method in EXPERIMENTAL_METHODS_BASE
    if PYINFORM_AVAILABLE or not method.startswith("te_")
]

##############################################
# Предобработка и загрузка
##############################################
def additional_preprocessing(df: pd.DataFrame, unique_thresh: float = 0.05) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        if len(df[col]) > 0 and pd.api.types.is_numeric_dtype(df[col]):
            uniq_ratio = df[col].nunique() / len(df[col])
            if uniq_ratio < unique_thresh:
                logging.info(f"[Preproc] Столбец {col} почти константный (uniq_ratio={uniq_ratio:.3f}), удаляем.")
                df.drop(columns=[col], inplace=True)
    for col in df.columns:
         if pd.api.types.is_numeric_dtype(df[col]) and (df[col] > 0).all():
            skew_before = stats.skew(df[col].dropna())
            if not np.isnan(skew_before):
                transformed = np.log(df[col])
                skew_after = stats.skew(transformed.dropna())
                if not np.isnan(skew_after) and abs(skew_after) < abs(skew_before):
                    logging.info(f"[Preproc] Лог-преобразование для {col}: skew {skew_before:.3f} -> {skew_after:.3f}.")
                    df[col] = transformed
    return df

def load_or_generate(filepath: str, log_transform=False, remove_outliers=True, normalize=True, fill_missing=True, check_stationarity=False) -> pd.DataFrame:
    try:
        if filepath.lower().endswith(".csv"):
            df = pd.read_csv(filepath, header=None)
        else:
            df = pd.read_excel(filepath, header=None)
        
        if df.shape[1] == 1 and isinstance(df.iloc[0,0], str):
            df = df[0].str.split('[,;]', expand=True)
        df = df.apply(pd.to_numeric, errors='coerce')

        if df.shape[0] < df.shape[1]:
            df = df.T
        
        num_cols = df.shape[1]
        df.columns = [f'c{i+1}' for i in range(num_cols)]
        
        df = df.fillna(df.mean())
        df = additional_preprocessing(df)
        df = df.copy()
        if log_transform:
            df = df.applymap(lambda x: np.log(x) if x is not None and not np.isnan(x) and x > 0 else x)
        if remove_outliers:
            for col in df.columns:
                if pd.api.types.is_numeric_dtype(df[col]):
                    series = df[col]
                    mean, std = series.mean(skipna=True), series.std(skipna=True)
                    if std > 0:
                        upper, lower = mean + DEFAULT_OUTLIER_Z * std, mean - DEFAULT_OUTLIER_Z * std
                        outliers = (series < lower) | (series > upper)
                        if outliers.any(): df.loc[outliers, col] = np.nan
        if fill_missing:
            df = df.interpolate(method='linear', limit_direction='both', axis=0).fillna(method='bfill').fillna(method='ffill').fillna(0)
        if normalize:
            cols_to_norm = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
            if cols_to_norm:
                scaler = StandardScaler()
                df[cols_to_norm] = scaler.fit_transform(df[cols_to_norm])
        if check_stationarity:
            for col in df.columns:
                 if pd.api.types.is_numeric_dtype(df[col]):
                    series = df[col].dropna()
                    if len(series) > 10:
                        pvalue = adfuller(series, autolag='AIC')[1]
                        logging.info(f"Ряд '{col}' {'стационарен' if pvalue <= 0.05 else 'вероятно нестационарен'} (p-value ADF={pvalue:.3f}).")
        
        logging.info(f"[Load] Данные успешно загружены, shape = {df.shape}.")
        return df
    except Exception as e:
        logging.error(f"[Load] Ошибка загрузки: {e}")
        raise e

##############################################
# Функции-метрики
##############################################
def correlation_matrix(data: pd.DataFrame, **kwargs) -> np.ndarray:
    return data.corr().values

def partial_correlation_matrix(df: pd.DataFrame, control: list = None, **kwargs) -> np.ndarray:
    cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    n = len(cols)
    out = np.eye(n)
    for i in range(n):
        for j in range(i + 1, n):
            xi, xj = cols[i], cols[j]
            ctrl_vars = control if control is not None else [c for c in cols if c not in (xi, xj)]
            sub_cols = [xi, xj] + [c for c in ctrl_vars if c in cols and c not in (xi, xj)]
            sub = df[sub_cols].dropna()
            if sub.shape[0] < len(sub_cols) + 1:
                pcor = np.nan
            else:
                try:
                    R = sub.corr().values
                    P = np.linalg.pinv(R)
                    pcor = -P[0, 1] / np.sqrt(P[0, 0] * P[1, 1])
                except Exception:
                    pcor = np.nan
            out[i, j] = out[j, i] = pcor
    return out


def partial_h2_matrix(df: pd.DataFrame, control: list = None, **kwargs) -> np.ndarray:
    """Возвращает квадрат частичной корреляции (H^2) для заданного контроля."""
    pcor = partial_correlation_matrix(df, control=control, **kwargs)
    return pcor**2

def lagged_directed_correlation(df: pd.DataFrame, lag: int, **kwargs) -> np.ndarray:
    m = df.shape[1]
    out = np.zeros((m, m))
    for i in range(m):
        for j in range(m):
            if i == j or len(df) <= lag: continue
            s1, s2 = df.iloc[:-lag, i].values, df.iloc[lag:, j].values
            if len(s1) > 1:
                out[i, j] = np.corrcoef(s1, s2)[0, 1]
    return out

def h2_matrix(df: pd.DataFrame, **kwargs) -> np.ndarray: return correlation_matrix(df)**2
def lagged_directed_h2(df: pd.DataFrame, lag: int, **kwargs) -> np.ndarray: return lagged_directed_correlation(df, lag)**2

def coherence_matrix(data: pd.DataFrame, **kwargs):
    N = data.shape[1]
    coh = np.zeros((N,N))
    for i in range(N):
        for j in range(i+1, N):
            s1, s2 = data.iloc[:,i].dropna(), data.iloc[:,j].dropna()
            if len(s1) > 1 and len(s2) > 1:
                try:
                    nperseg = min(256, min(len(s1), len(s2)))
                    f, Cxy = signal.coherence(s1, s2, nperseg=nperseg)
                    coh[i,j] = coh[j,i] = float(np.nanmean(Cxy))
                except ValueError:
                    coh[i,j] = coh[j,i] = np.nan
    return coh


def _knn_entropy(X, k=DEFAULT_K_MI):
    """Вычисляет энтропию для 1D-массива с помощью KNN."""
    N = len(X)
    if N <= k: return 0.0
    tree = cKDTree(X.reshape(-1, 1))
    d, _ = tree.query(X.reshape(-1, 1), k=k + 1, p=np.inf)
    # Расстояние до k-го соседа
    r = d[:, k]
    # digamma(N) - digamma(k) + d*log(2*r_k) - это для d-мерного пространства
    # Для 1D: digamma(N) - digamma(k) + E[log(2r_k)]
    return digamma(N) - digamma(k) + np.mean(np.log(2 * r + 1e-10)) 


def _knn_mutual_info(X, Y, k=DEFAULT_K_MI):
    """
    Вычисляет 'сырую' взаимную информацию (MI) с помощью KNN.
    """
    N = len(X)
    if N <= k: return 0.0
    
    XY = np.c_[X, Y]
    tree_XY = cKDTree(XY)
    d, _ = tree_XY.query(XY, k=k+1, p=np.inf)
    eps = d[:,k]
    nx = np.array([len(cKDTree(X.reshape(-1,1)).query_ball_point(X[i:i+1], r=e, p=np.inf)) for i, e in enumerate(eps)])
    ny = np.array([len(cKDTree(Y.reshape(-1,1)).query_ball_point(Y[i:i+1], r=e, p=np.inf)) for i, e in enumerate(eps)])
    raw_mi = digamma(N) + digamma(k) - np.mean(digamma(nx) + digamma(ny))
    
    return max(0, raw_mi)


def mutual_info_matrix(data: pd.DataFrame, k=DEFAULT_K_MI, **kwargs):
    """
    Строит матрицу 'СЫРОЙ' взаимной информации (MI)
    """
    n_vars = data.shape[1]
    mi = np.zeros((n_vars, n_vars))
    
    for i in range(n_vars):
        for j in range(i + 1, n_vars):
            s1 = data.iloc[:, i].dropna().values
            s2 = data.iloc[:, j].dropna().values
            
            min_len = min(len(s1), len(s2))
            if min_len > k:
                val = _knn_mutual_info(s1[:min_len], s2[:min_len], k=k)
                mi[i, j] = mi[j, i] = val
                
    return mi
def _knn_conditional_mutual_info(X, Y, Z, k=DEFAULT_K_MI):
    N = len(X)
    if N <= k: return 0.0
    XZ, YZ = np.c_[X,Z], np.c_[Y,Z]
    tree_XZ, tree_YZ, tree_Z = cKDTree(XZ), cKDTree(YZ), cKDTree(Z.reshape(N,-1))
    d, _ = cKDTree(np.c_[X,Y,Z]).query(np.c_[X,Y,Z], k=k+1, p=np.inf)
    eps = d[:,k]
    nxz = np.array([len(tree_XZ.query_ball_point(XZ[i], r=e, p=np.inf)) for i, e in enumerate(eps)])
    nyz = np.array([len(tree_YZ.query_ball_point(YZ[i], r=e, p=np.inf)) for i, e in enumerate(eps)])
    nz = np.array([len(tree_Z.query_ball_point(Z[i:i+1].reshape(-1), r=e, p=np.inf)) for i, e in enumerate(eps)])
    return digamma(k) - np.mean(digamma(nxz) + digamma(nyz) - digamma(nz))

def mutual_info_matrix_partial(data: pd.DataFrame, control: Optional[List[str]] = None, k=DEFAULT_K_MI, **kwargs):
    cols = list(data.columns)
    N = len(cols)
    pmi = np.zeros((N, N))
    for i in range(N):
        for j in range(i+1, N):
            xi, xj = cols[i], cols[j]
            Z_cols = control if control is not None else [c for c in cols if c not in (xi, xj)]
            Z_cols = [c for c in Z_cols if c in data.columns and c not in (xi, xj)]
            
            if not Z_cols:
                s1, s2 = data[xi].dropna(), data[xj].dropna()
                min_len = min(len(s1), len(s2))
                val = _knn_mutual_info(s1.values[:min_len], s2.values[:min_len], k=k)
            else:
                sub = data[[xi, xj] + Z_cols].dropna()
                if len(sub) <= k: val = np.nan
                else:
                    X, Y, Z = sub[xi].values, sub[xj].values, sub[Z_cols].values
                    val = _knn_conditional_mutual_info(X, Y, Z, k=k)
            pmi[i,j] = pmi[j,i] = max(0, val) if not np.isnan(val) else np.nan
    return pmi

def compute_granger_matrix(df: pd.DataFrame, lags: int = DEFAULT_MAX_LAG, **kwargs) -> np.ndarray:
    n = df.shape[1]
    G = np.full((n, n), 1.0)
    cols = df.columns.tolist()
    for i in range(n):
        for j in range(n):
            if i == j: continue
            data_pair = df[[cols[j], cols[i]]].dropna()
            if len(data_pair) > lags * 2 + 5:
                try:
                    tests = grangercausalitytests(data_pair, maxlag=[lags], verbose=False)
                    G[i, j] = tests[lags][0]['ssr_ftest'][1]
                except (np.linalg.LinAlgError, ValueError): 
                    G[i, j] = np.nan
    return G

def _load_pyinform():
    """Ленивая загрузка pyinform без падения всего приложения."""
    if not PYINFORM_AVAILABLE:
        logging.warning(
            "[PyInform] pyinform не установлен: TE-методы будут отключены.",
        )
        return None
    return importlib.import_module("pyinform")


def compute_TE(source: np.ndarray, target: np.ndarray, lag: int = 1, bins: int = DEFAULT_BINS):
    """Transfer Entropy с защитой от отсутствия pyinform."""
    try:
        pyinform = _load_pyinform()
        if pyinform is None:
            return np.nan

        def discretize_manual(series, num_bins):
            series_f = series.astype(np.float64)
            min_val, max_val = np.min(series_f), np.max(series_f)
            if min_val == max_val: 
                return np.zeros(len(series_f), dtype=int)
            bin_edges = np.linspace(min_val, max_val, num=num_bins + 1)
            bin_edges[-1] += 1e-9
            discretized = np.digitize(series_f, bins=bin_edges) - 1
            return discretized

        source_discrete = discretize_manual(source, bins)
        target_discrete = discretize_manual(target, bins)

        te = pyinform.transfer_entropy(source_discrete, target_discrete, k=lag)
        return te
    except Exception as e:
        logging.error(f"[PyInform TE] Ошибка вычисления: {e}")
        return np.nan

def TE_matrix(df: pd.DataFrame, lag: int = 1, bins: int = DEFAULT_BINS, **kwargs):
    """
    Строит матрицу Transfer Entropy для всех пар
    """
    n = df.shape[1]
    te_matrix = np.zeros((n, n))
    
    for i in range(n):
        for j in range(n): 
            if i == j:
                continue
            
            s1 = df.iloc[:, i].dropna().values
            s2 = df.iloc[:, j].dropna().values
            
            min_len = min(len(s1), len(s2))
            if min_len <= lag:
                te_matrix[j, i] = np.nan
                continue
            
            s1 = s1[:min_len]
            s2 = s2[:min_len]
            te_matrix[j, i] = compute_TE(s1, s2, lag=lag, bins=bins)
            
    return te_matrix

def TE_matrix_partial(df: pd.DataFrame,
                      lag: int = 1,
                      control: Optional[List[str]] = None,
                      bins: int = DEFAULT_BINS) -> np.ndarray:
    cols = list(df.columns)
    N = len(cols)
    M = np.zeros((N, N))

    def discrete_conditional_mutual_info(X, Y, Z, bins):
        return np.nan 

    for i, src in enumerate(cols):
        for j, tgt in enumerate(cols):
            if i == j:
                continue

            X_series = df[src].shift(lag)
            Y_series = df[tgt]

            Z_parts = []
            Z_parts.append(df[tgt].shift(lag))
            ctrl = control if control is not None else [c for c in cols if c not in (src, tgt)]
            for c in ctrl:
                if c in df.columns:
                    Z_parts.append(df[c].shift(lag))
            
            combined_df = pd.concat([X_series.rename('X'), Y_series.rename('Y')] + Z_parts, axis=1).dropna()

            if len(combined_df) < 2:
                M[i, j] = np.nan 
                continue

            Xv = combined_df['X'].values
            Yv = combined_df['Y'].values
            Zv = combined_df.iloc[:, 2:].values
            
            try:
                M[i, j] = discrete_conditional_mutual_info(Xv, Yv, Zv, bins=bins)
            except Exception:
                M[i, j] = np.nan 

    return M


def AH_matrix(df: pd.DataFrame, embed_dim=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU) -> np.ndarray:
    df = df.dropna(axis=0, how='any')
    N = df.shape[1]
    AH_matrix = np.zeros((N, N))
    arr = df.values
    for i in range(N):
        for j in range(N):
            if i == j:
                AH_matrix[i, j] = 0.0
            else:
                H_val = _H_ratio_direction(arr[:, j], arr[:, i], m=embed_dim, tau=tau)
                if H_val is None or H_val <= 0:
                    AH = 0.0
                else:
                    AH = 1.0 / H_val
                    if AH > 1.0: 
                        AH = 1.0
                AH_matrix[i, j] = AH
    return AH_matrix

def _H_ratio_direction(X, Y, m=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU):
    n = len(X)
    if len(Y) != n or n < 2:
        return None
    L = n - (m - 1) * tau
    if L < 2:
        return None
    
    X_state = np.zeros((L, m))
    Y_state = np.zeros((L, m))
    for j in range(m):
        X_state[:, j] = X[j*tau : j*tau + L]
        Y_state[:, j] = Y[j*tau : j*tau + L]
    
    valid_indices = ~np.isnan(X_state).any(axis=1) & ~np.isnan(Y_state).any(axis=1)
    if not np.any(valid_indices):
        return None
        
    X_state_valid = X_state[valid_indices]
    Y_state_valid = Y_state[valid_indices]
    
    if len(X_state_valid) < 2:
        return None

    tree_X = cKDTree(X_state_valid)
    tree_Y = cKDTree(Y_state_valid)
    
    dists_X, idx_X = tree_X.query(X_state_valid, k=2)
    
    
    if idx_X.shape[1] < 2: 
        return None 
    
    nn_idx = idx_X[:, 1] 
    diff = Y_state_valid - Y_state_valid[nn_idx]
    dY1 = np.sqrt(np.sum(diff**2, axis=1))
    
    dists_Y, _ = tree_Y.query(Y_state_valid, k=2)
    dY2 = dists_Y[:, 1]
    dY2 = np.where(dY2 == 0, 1e-10, dY2) 
    
    ratios = dY1 / dY2
    ratios = ratios[np.isfinite(ratios)] 
    
    if len(ratios) == 0:
        return None

    H_val = np.mean(ratios)
    return H_val

def compute_partial_AH_matrix(data: pd.DataFrame,
                               max_lag: int = DEFAULT_MAX_LAG,
                               embed_dim: int = DEFAULT_EMBED_DIM,
                               tau: int = DEFAULT_EMBED_TAU,
                               control: List[str] = None) -> np.ndarray:
    df = data.dropna(axis=0, how='any')
    N = df.shape[1]
    if N < 2:
        return np.zeros((N, N))

    if control and len(control) > 0:
        resid_df = pd.DataFrame(index=df.index)
        for col in df.columns:
            X_ctrl = df[control]
            y = df[col]
            if len(X_ctrl) > 0 and len(y) == len(X_ctrl) and not X_ctrl.isnull().any().any():
                try:
                    model = LinearRegression().fit(X_ctrl.values, y.values)
                    resid = y.values - model.predict(X_ctrl.values)
                    resid_df[col] = resid
                except ValueError: 
                    resid_df[col] = y 
            else:
                resid_df[col] = y
    else:
        try:
            model = VAR(df.values)
            res_full = model.fit(max_lag, ic=None)
            resid_df = pd.DataFrame(res_full.resid, columns=df.columns)
        except Exception as e:
            logging.error(f"VAR fit error (partial AH, fallback to raw): {e}")
            resid_df = df 

    return AH_matrix(resid_df, embed_dim=embed_dim, tau=tau)


def directional_AH_matrix(df: pd.DataFrame, maxlags: int = 5) -> np.ndarray:
    return AH_matrix(df, embed_dim=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU)

def granger_dict(df: pd.DataFrame, maxlag: int = 4) -> dict:
    results = {}
    cols = list(df.columns)
    for i, tgt in enumerate(cols):
        for j, src in enumerate(cols):
            if i == j:
                continue
            sub = df[[tgt, src]].dropna()
            if len(sub) < (maxlag + 10):
                results[f"{src}->{tgt}"] = None
                continue
            try:
                tests = grangercausalitytests(sub, maxlag=maxlag, verbose=False)
            except Exception as e:
                logging.error(f"[Granger] Ошибка Granger для {src}->{tgt}: {e}")
                results[f"{src}->{tgt}"] = None
                continue 
            results[f"{src}->{tgt}"] = tests # Сохраняем РЕЗУЬТАТ
    return results

# эта матрица НЕ ТА ЖЕ ЧТО В МАППИНГЕ
def _compute_granger_matrix_internal(df: pd.DataFrame, lags: int = DEFAULT_MAX_LAG) -> np.ndarray:
    n = df.shape[1]
    G = np.zeros((n, n))
    cols = df.columns.tolist()
    for i in range(n):
        for j in range(n):
            if i == j:
                G[i, j] = 0.0
            else:
                sub = df[[cols[j], cols[i]]].dropna()
                try:
                    tests = grangercausalitytests(sub, maxlag=[lags], verbose=False)
                    pvals = [tests[l][0]['ssr_ftest'][1] for l in tests]
                    G[i, j] = min(pvals)
                except Exception as e: 
                    logging.error(f"[Granger-Internal] Ошибка Granger для {cols[j]}->{cols[i]}: {e}")
                    G[i, j] = np.nan
    return G


def compute_partial_granger_matrix(data: pd.DataFrame, lags=DEFAULT_MAX_LAG) -> np.ndarray:
    """
    контроль остальных переменных, грейндж
    """
    df = data.dropna(axis=0, how='any')
    N = df.shape[1]
    if N < 3:
        return _compute_granger_matrix_internal(data, lags=lags)
    pg_matrix = np.zeros((N, N))
    T = len(df)
    p = lags
    if T <= p:
        return pg_matrix
    arr = df.values
    try:
        model_full = VAR(arr)
        res_full = model_full.fit(p, ic=None)
    except Exception as e:
        logging.error(f"VAR fit error (partial Granger): {e}")
        return pg_matrix
    sigma_full = np.cov(res_full.resid, rowvar=False)
    for i in range(N):
        reduced_arr = np.delete(arr, i, axis=1)
        try:
            model_red = VAR(reduced_arr)
            res_red = model_red.fit(p, ic=None)
            sigma_red = np.cov(res_red.resid, rowvar=False)
        except Exception as e:
            for j in range(N):
                if j != i:
                    pg_matrix[i, j] = np.nan 
            continue
        for j in range(N):
            if i == j:
                pg_matrix[i, j] = 0.0
            else:
                idx_j = j - 1 if i < j else j
                var_full = sigma_full[j, j] if sigma_full.shape[0] > j else np.var(res_full.resid[:, j])
                var_red = sigma_red[idx_j, idx_j] if sigma_red.shape[0] > idx_j else np.var(res_red.resid[:, idx_j])
                if var_full <= 0 or var_red <= 0:
                    gc_val = np.nan 
                else:
                    gc_val = np.log(var_red / var_full)
                    if gc_val < 0:
                        gc_val = 0.0
                pg_matrix[i, j] = gc_val
    return pg_matrix

def granger_matrix(df: pd.DataFrame, granger_dict_result: dict) -> np.ndarray:
    cols = list(df.columns)
    n_vars = len(cols)
    G = np.ones((n_vars, n_vars))
    for i, tgt in enumerate(cols):
        for j, src in enumerate(cols):
            if i == j:
                G[i, j] = 0
            else:
                key = f"{src}->{tgt}"
                if granger_dict_result.get(key) is None:
                    G[i, j] = np.nan
                else:
                    test_dict = granger_dict_result[key]
                    bp = 1.0 
                    found_valid_p = False 
                    for lag_val, dct in test_dict.items():
                        if isinstance(dct, list) and len(dct) > 0 and 'ssr_ftest' in dct[0]:
                            F, pval, _, _ = dct[0]['ssr_ftest']
                            if not np.isnan(pval): 
                                bp = min(bp, pval)
                                found_valid_p = True
                    G[i, j] = bp if found_valid_p else np.nan 
    return G

def granger_matrix_partial(df: pd.DataFrame, maxlag: int = 4, control: list = None) -> np.ndarray:
    columns = list(df.columns)
    n_vars = len(columns)
    G = np.ones((n_vars, n_vars))
    
    def remove_linear_dependency(df, src, tgt, control_cols):
        #НАДО РЕАЛИЗОВАТЬ ЗАПОЛНИТЕЛЬ НОРМАЛЬНО, удалять и не добавлять НЕХОРОШО
        if not control_cols:
            return df[src].values, df[tgt].values
        
        X_control = df[control_cols].values
        y_src = df[src].values
        y_tgt = df[tgt].values
        
        try:
            reg_src = LinearRegression().fit(X_control, y_src)
            resid_src = y_src - reg_src.predict(X_control)
            
            reg_tgt = LinearRegression().fit(X_control, y_tgt)
            resid_tgt = y_tgt - reg_tgt.predict(X_control)
            return resid_src, resid_tgt
        except Exception:
            return np.array([]), np.array([]) #фейлится если матрица плохая


    for i, tgt in enumerate(columns):
        for j, src in enumerate(columns):
            if i == j:
                G[i, j] = 0
            else:
                control_cols = control if control is not None else [c for c in columns if c not in [src, tgt]]
                sub = df[[src, tgt] + control_cols].dropna()
                if len(sub) < (maxlag + 10):
                    G[i, j] = np.nan
                    continue
                r1, r2 = remove_linear_dependency(sub, src, tgt, control_cols)
                if r1.size == 0 or r2.size == 0 or len(r1) != len(r2):
                    G[i, j] = np.nan
                    continue
                dmini = pd.DataFrame({'yy': r2, 'xx': r1}).dropna()
                if len(dmini) < (maxlag + 10):
                    G[i, j] = np.nan
                    continue
                try:
                    gg = grangercausalitytests(dmini[['yy', 'xx']], maxlag=maxlag, verbose=False)
                    bp = 1.0
                    for lag, dct in gg.items():
                        F, pval, _, _ = dct[0]['ssr_ftest']
                        if pval < bp:
                            bp = pval
                    G[i, j] = bp
                except Exception as e:
                    logging.error(f"[Granger-partial] Ошибка Granger (partial) для {src}->{tgt}: {e}")
                    G[i, j] = np.nan
    return G

##############################################
# Функции для частотного анализа
##############################################
def plt_fft_analysis(series: pd.Series):
    arr = series.dropna().values
    if len(arr) == 0:
        return np.array([]), np.array([]), np.array([]), np.array([])
    n = len(arr)
    dt = 10 / (n - 1) if n > 1 else 1.0
    freqs = np.fft.fftfreq(n, d=dt)
    fft_vals = fft(arr)
    amplitude = np.abs(fft_vals)
    phase = np.angle(fft_vals)
    pos_mask = freqs >= 0
    freqs, amplitude, phase = freqs[pos_mask], amplitude[pos_mask], phase[pos_mask]
    peaks, _ = find_peaks(amplitude, height=(np.max(amplitude)*0.2 if amplitude.size > 0 else 0))
    logging.debug(f"[FFT] Найдено пиков на частотах: {freqs[peaks] if peaks.size > 0 else 'Нет пиков'}")
    return freqs, amplitude, phase, peaks

def plot_amplitude_response(series: pd.Series, title: str) -> BytesIO:
    freqs, amplitude, phase, peaks = plt_fft_analysis(series)
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, amplitude, label="АЧХ")
    if peaks.size > 0:
        ax.plot(freqs[peaks], amplitude[peaks], "x", label="Пики")
    ax.set_title(title)
    ax.set_xlabel("Частота")
    ax.set_ylabel("Амплитуда")
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

def plot_phase_response(series: pd.Series, title: str) -> BytesIO:
    freqs, amplitude, phase, peaks = plt_fft_analysis(series)
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, phase, label="ФЧХ", color="orange")
    ax.set_title(title)
    ax.set_xlabel("Частота")
    ax.set_ylabel("Фаза (рад)")
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

def plot_combined_ac_fch(data: pd.DataFrame, title: str) -> BytesIO:
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 8), sharex=True)
    for col in data.columns:
        series = data[col]
        freqs, amplitude, phase, _ = plt_fft_analysis(series)
        if freqs.size > 0:
            ax1.plot(freqs, amplitude, label=col)
            ax2.plot(freqs, phase, label=col)
    ax1.set_title(title + " - АЧХ")
    ax1.set_ylabel("Амплитуда")
    ax1.legend()
    ax2.set_title(title + " - ФЧХ")
    ax2.set_xlabel("Частота")
    ax2.set_ylabel("Фаза (рад)")
    ax2.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

##############################################
# индивидуальный AC & PH для каждого ряда
##############################################
def plot_individual_ac_ph(data: pd.DataFrame, title: str) -> dict:
    plots = {}
    for col in data.columns:
        series = data[col]
        # График АЧХ
        fig1, ax1 = plt.subplots(figsize=(6, 4))
        freqs, amplitude, _, peaks = plt_fft_analysis(series)
        ax1.plot(freqs, amplitude, label=f"АЧХ {col}")
        if peaks.size > 0:
            ax1.plot(freqs[peaks], amplitude[peaks], "x", label="Пики")
        ax1.set_title(f"АЧХ {col}")
        ax1.set_xlabel("Частота")
        ax1.set_ylabel("Амплитуда")
        ax1.legend()
        buf1 = BytesIO()
        plt.tight_layout()
        plt.savefig(buf1, format="png", dpi=100)
        buf1.seek(0)
        plt.close(fig1)
        # График ФЧХ
        fig2, ax2 = plt.subplots(figsize=(6, 4))
        freqs, _, phase, _ = plt_fft_analysis(series)
        ax2.plot(freqs, phase, label=f"ФЧХ {col}", color="orange")
        ax2.set_title(f"ФЧХ {col}")
        ax2.set_xlabel("Частота")
        ax2.set_ylabel("Фаза (рад)")
        ax2.legend()
        buf2 = BytesIO()
        plt.tight_layout()
        plt.savefig(buf2, format="png", dpi=100)
        buf2.seek(0)
        plt.close(fig2)
        plots[col] = {"AC": buf1, "PH": buf2}
    return plots

##############################################
# sample entropy
##############################################
def compute_sample_entropy(series: pd.Series) -> float:
    try:
        return nolds.sampen(series.dropna().values)
    except Exception as ex:
        logging.error(f"[Sample Entropy] Ошибка: {ex}")
        return np.nan

##############################################
# Функции для частотного анализа когерентности
##############################################
def plot_coherence_vs_frequency(series1: pd.Series, series2: pd.Series, title: str, fs: int = 100) -> BytesIO:
    s1 = series1.dropna().values
    s2 = series2.dropna().values
    n = min(len(s1), len(s2))
    if n == 0:
        return BytesIO()
    s1 = s1[:n]
    s2 = s2[:n]
    freqs, cxy = coherence(s1, s2, fs=fs)
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, cxy, label="Когерентность")
    if cxy.size > 0:
        max_idx = np.argmax(cxy)
        max_freq = freqs[max_idx]
        max_coh = cxy[max_idx]
        ax.plot(max_freq, max_coh, "ro", label=f"Макс. связь: {max_coh:.3f} на {max_freq:.3f}Hz")
        ax.annotate(f"{max_freq:.3f} Hz", xy=(max_freq, max_coh), xytext=(max_freq, max_coh+0.05),
                    arrowprops=dict(facecolor='black', shrink=0.05))
    ax.set_title(title)
    ax.set_xlabel("Частота (Hz)")
    ax.set_ylabel("Когерентность")
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)
    return buf

##############################################
# Функции для экспорта данных в Excel
##############################################
def add_raw_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Добавляет лист с исходными данными."""
    ws = wb.create_sheet("Raw Data")
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))

def plot_heatmap(matrix: np.ndarray, title: str, legend_text: str = "", annotate: bool = False, vmin=None, vmax=None) -> BytesIO:
    fig, ax = plt.subplots(figsize=(4, 3.2))
    
    if matrix is None or not isinstance(matrix, np.ndarray) or matrix.size == 0:
        ax.text(0.5, 0.5, "Error\n(No Data)", ha='center', va='center', color='red', fontsize=12)
        ax.set_xticks([])
        ax.set_yticks([])
    else:
        # Фиксируем шкалу, чтобы избежать авто-нормализации matplotlib.
        cax = ax.imshow(matrix, cmap="viridis", aspect="auto", vmin=vmin, vmax=vmax)
        fig.colorbar(cax, ax=ax)
        ax.set_title(title, fontsize=10)
        
        # Аннотации
        if annotate and matrix.shape[0] < 10:
            min_val = vmin if vmin is not None else np.nanmin(matrix)
            max_val = vmax if vmax is not None else np.nanmax(matrix)
            
            if np.isfinite(min_val) and np.isfinite(max_val) and max_val > min_val:
                threshold = min_val + (max_val - min_val) / 2.0
            else:
                threshold = 0.5 # запасной вариант

            for i in range(matrix.shape[0]):
                for j in range(matrix.shape[1]):
                    val = matrix[i, j]
                    if np.isnan(val):
                        display_val, color = "NaN", "red"
                    else:
                        display_val = f"{val:.2f}"
                        # Цвет текста зависит от порога, который выч отдельно. от 0,5 НЕ ВОЗВРАЩАТЬ
                        color = "white" if val < threshold else "black"
                    ax.text(j, i, display_val, ha="center", va="center", color=color, fontsize=8)

    if legend_text:
        ax.text(0.05, 0.95, legend_text, transform=ax.transAxes, fontsize=8,
                verticalalignment='top', bbox=dict(facecolor='white', alpha=0.5))
    
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=150)
    buf.seek(0)
    plt.close(fig)
    return buf

def plot_connectome(matrix: np.ndarray, method_name: str, threshold: float = 0.2,
                    directed: bool = False, invert_threshold: bool = False, legend_text: str = "") -> BytesIO:
    n = matrix.shape[0]
    G = nx.DiGraph() if directed else nx.Graph()
    G.add_nodes_from(range(n))
    for i in range(n):
        for j in range(n):
            if i != j and matrix[i, j] is not None and not np.isnan(matrix[i, j]):
                if invert_threshold:
                    if matrix[i, j] < threshold:
                        G.add_edge(j, i, weight=matrix[i, j])
                else:
                    if abs(matrix[i, j]) > threshold:
                        G.add_edge(j, i, weight=matrix[i, j])
    pos = nx.circular_layout(G)
    fig, ax = plt.subplots(figsize=(4, 4))
    if directed:
        nx.draw_networkx_nodes(G, pos, ax=ax, node_color="lightblue", node_size=500)
        nx.draw_networkx_labels(G, pos, ax=ax)
        nx.draw_networkx_edges(G, pos, ax=ax, arrowstyle="->", arrowsize=10)
    else:
        nx.draw_networkx(G, pos, ax=ax, node_color="lightblue", node_size=500)
    ax.set_title(f"Connectome: {method_name}")
    if legend_text:
         ax.text(0.05, 0.05, legend_text, transform=ax.transAxes, fontsize=8,
                 verticalalignment='bottom', bbox=dict(facecolor='white', alpha=0.5))
    ax.axis("off")
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)
    return buf

def add_method_to_sheet(ws, row: int, title: str, matrix: np.ndarray, directed: bool = False, legend_text: str = "") -> int:
    ws.append([title])
    if matrix is None:
        ws.append(["Метод не работает для этих данных."])
        return ws.max_row
    df_mat = pd.DataFrame(matrix)
    for r in dataframe_to_rows(df_mat, index=False, header=True):
        ws.append(r)
    buf_heat = plot_heatmap(matrix, title + " Heatmap", legend_text=legend_text)
    img_heat = Image(buf_heat)
    img_heat.width = 400
    img_heat.height = 300
    ws.add_image(img_heat, f"A{ws.max_row + 2}")
    buf_conn = plot_connectome(matrix, title + " Connectome", threshold=0.2, directed=directed, invert_threshold=False, legend_text=legend_text)
    img_conn = Image(buf_conn)
    img_conn.width = 400
    img_conn.height = 400
    ws.add_image(img_conn, f"G{ws.max_row + 2}")
    return ws.max_row

def fmt_val(v):
    try:
        f = float(v)
        if np.isnan(f):
            return "N/A"
        return f"{f:.3f}"
    except Exception:
        return "N/A"
#
# МАППИНГ
###
method_mapping = {
    # ——— Correlation ———
    "correlation_full":     lambda data, lag=None, control=None: correlation_matrix(data),
    "correlation_partial":  lambda data, lag=None, control=None: partial_correlation_matrix(data, control),
    "correlation_directed": lambda data, lag, control=None: lagged_directed_correlation(data, lag),

    # ——— H² (squared corr) ———
    "h2_full":              lambda data, lag=None, control=None: correlation_matrix(data)**2,
    "h2_partial":           lambda data, lag=None, control=None: partial_h2_matrix(data, control), 
    "h2_directed":          lambda data, lag, control=None: lagged_directed_h2(data, lag),

    # ——— Mutual Information ———
    "mutinf_full":          lambda data, lag=0, control=None: mutual_info_matrix(data, k=DEFAULT_K_MI),
    "mutinf_partial":       lambda data, lag=0, control=None: mutual_info_matrix_partial(data, control, k=DEFAULT_K_MI),

    # ——— Coherence ———
    "coherence_full":       lambda data, lag=None, control=None: coherence_matrix(data),


# ...
    # ——— Granger causality ———
    "granger_full":         lambda data, lag, control=None: _compute_granger_matrix_internal(data, lags=lag),
    "granger_partial":      lambda data, lag, control=None: compute_partial_granger_matrix(data, lags=lag), 
    "granger_directed":     lambda data, lag, control=None: _compute_granger_matrix_internal(data, lags=lag),
# ...

    # ——— Transfer entropy ———
    "te_full":    lambda data, lag, control=None: TE_matrix(data, lag=lag),
    "te_partial": lambda data, lag, control=None: TE_matrix_partial(data, lag=lag, control=control, bins=DEFAULT_BINS),
    "te_directed":lambda data, lag, control=None: TE_matrix_partial(data, lag=lag, control=control, bins=DEFAULT_BINS),


    # ——— AH (non‑linear) ———
    "ah_full":              lambda data, lag=None, control=None: AH_matrix(data),
    "ah_partial":           lambda data, lag, control=None: compute_partial_AH_matrix(data, max_lag=lag, control=control),
    "ah_directed":          lambda data, lag, control=None:
                                (AH_matrix(data)
                                 if not control
                                 else compute_partial_AH_matrix(data, max_lag=lag, control=control)),
}

def compute_connectivity_variant(data, variant, lag=1, control=None):
    try:
        if control is not None and len(control) == 0:
            control = None
        # если есть метод в mapping
        if variant in method_mapping:
            return method_mapping[variant](data, lag, control)
        # иначе — корреляция
        return correlation_matrix(data)
    except Exception as e:
        logging.error(f"[ComputeVariant] Метод {variant} не работает: {e}")
        return None


def powerset(iterable):
    s = list(iterable)
    return chain.from_iterable(combinations(s, r) for r in range(len(s)+1))

##############################################
#  диагностика коэффициентов регрессии
##############################################
def regression_diagnostics(df: pd.DataFrame, target: str, controls: list):
    """
    Рассчитывает линейную регрессию вида: target ~ controls.
    Возвращает строку с R².
    """
    # Если нет контрольных переменных — выходим
    if not controls:
        return f"Нет контрольных переменных для {target}."
    # Иначе строим модель и возвращаем R²
    X = df[controls]
    y = df[target]
    model = LinearRegression().fit(X, y)
    r2 = model.score(X, y)
    return f"{target} ~ {controls}: R² = {r2:.3f}"


##############################################
# Частотный анализ: возвращает пиковые значения
##############################################
def frequency_analysis(series: pd.Series, peak_height_ratio: float = 0.2):
    freqs, amplitude, phase, peaks = plt_fft_analysis(series)
    if freqs.size == 0 or peaks.size == 0:
        return None, None, None
    peak_freqs = freqs[peaks]
    peak_amps = amplitude[peaks]
    periods = 1 / peak_freqs
    return peak_freqs, peak_amps, periods

def sliding_fft_analysis(data: pd.DataFrame, window_size: int, overlap: int) -> dict:
    """Экспериментальный анализ скользящего FFT (по умолчанию отключён)."""
    logging.info("[Sliding FFT] Экспериментальная функция отключена.")
    return {}


def analyze_sliding_windows_with_metric(
    data: pd.DataFrame,
    variant: str,
    window_size: int,
    overlap: int,
) -> dict:
    """Экспериментальный анализ скользящих окон (по умолчанию отключён)."""
    logging.info("[Sliding Window] Экспериментальная функция отключена.")
    return {}


def sliding_window_pairwise_analysis(
    data: pd.DataFrame,
    method: str,
    window_size: int,
    overlap: int,
) -> dict:
    """Экспериментальный парный анализ скользящих окон (по умолчанию отключён)."""
    logging.info("[Sliding Pairwise] Экспериментальная функция отключена.")
    return {}

##############################################
# Листы с коэффициентами и частотным анализом
##############################################
def export_coefficients_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Coefficients & Explanations")
    ws.append(["Описание:", "Лист содержит краткие пояснения коэффициентов регрессий и матриц связей."])
    ws.append(["Например, коэффициенты регрессии показывают, как контрольные переменные влияют на связь между переменными."])
    ws.append([])
    ws.append(["Регрессионная диагностика:"])
    ws.append(["Переменная", "Контроль", "Диагностика"])
    for target in tool.data.columns:
        controls = [c for c in tool.data.columns if c != target]
        diag_str = regression_diagnostics(tool.data, target, controls)
        ws.append([target, str(controls), diag_str])
    ws.append([])
    ws.append(["Матрицы связей:"])
    ws.append(["Метод", "Описание"])
    methods_info = [
        ("correlation_full", "Стандартная корреляционная матрица."),
        ("correlation_partial", "Частичная корреляция (с контролем)."),
        ("mutinf_full", "Полная взаимная информация."),
        ("coherence_full", "Когерентность между переменными.")
    ]
    for m, info in methods_info:
        ws.append([m, info])
    logging.info("[Coefficients] Лист 'Coefficients & Explanations' сформирован.")

def export_frequency_summary_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Frequency Summary")
    ws.append(["Столбец", "Пиковые частоты", "Пиковые амплитуды", "Периоды", "Пояснение"])
    for col in tool.data.columns:
        s = tool.data[col].dropna()
        freq, amps, periods = frequency_analysis(s)
        if freq is not None:
            freq_str = ", ".join([f"{f:.3f}" for f in freq])
            amps_str = ", ".join([f"{f:.3f}" for f in amps])
            period_str = ", ".join([f"{p:.1f}" for p in periods])
            note = f"Макс. связь на {freq[np.argmax(amps)]:.3f} Hz"
        else:
            freq_str = amps_str = period_str = "Нет пиков"
            note = "Пиковые частоты не выявлены"
        ws.append([col, freq_str, amps_str, period_str, note])
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
    logging.info("[Frequency] Лист 'Frequency Summary' сформирован.")

##############################################
# Новый лист: Индивидуальные АЧХ и ФЧХ (раздельно)
##############################################
def export_individual_ac_ph_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Individual AC & PH")
    ws.append(["Столбец", "АЧХ", "ФЧХ"])
    plots = plot_individual_ac_ph(tool.data_normalized, "Individual AC & PH")
    for col, imgs in plots.items():
        ws.append([col])
        img_ac = Image(imgs["AC"])
        img_ac.width = 400
        img_ac.height = 300
        ws.add_image(img_ac, f"B{ws.max_row}")
        img_ph = Image(imgs["PH"])
        img_ph.width = 400
        img_ph.height = 300
        ws.add_image(img_ph, f"G{ws.max_row}")
    logging.info("[Individual AC & PH] Лист сформирован.")

##############################################
# Новый лист: Анализ энтропии
##############################################
def export_entropy_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Entropy Analysis")
    ws.append(["Столбец", "Sample Entropy (sampen)"])
    for col in tool.data.columns:
        s = tool.data[col].dropna()
        ent = compute_sample_entropy(s)
        ws.append([col, f"{ent:.3f}" if not np.isnan(ent) else "N/A"])
    logging.info("[Entropy Analysis] Лист сформирован.")

##############################################
# Новый лист
##############################################
def export_combined_informational_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined Informational Analysis")
    current_row = 1
    ws.cell(row=current_row, column=1, value="Combined Informational Analysis")
    current_row += 2
    ws.cell(row=current_row, column=1, value="Lag Analysis Summary (Aggregated)")
    current_row += 1
    buf_lag = tool.plot_all_methods_lag_comparison(tool.lag_results)
    img_lag = Image(buf_lag)
    img_lag.width = 800
    img_lag.height = 600
    ws.add_image(img_lag, f"A{current_row}")
    current_row += 30
    ws.cell(row=current_row, column=1, value="Sliding Window Analysis Summary (Aggregated)")
    current_row += 1
    sw_res = tool.analyze_sliding_windows(
        "coherence_full",
        window_size=min(50, len(tool.data_normalized) // 2),
        overlap=min(25, len(tool.data_normalized) // 4),
    )
    if sw_res:
        legend_text = "Метод: coherence_full, Окно: 50"
        buf_sw = tool.plot_sliding_window_comparison(sw_res, legend_text=legend_text)
        img_sw = Image(buf_sw)
        img_sw.width = 700
        img_sw.height = 400
        ws.add_image(img_sw, f"A{current_row}")
        current_row += 20
    else:
        ws.append(["Sliding Window Analysis отключён или нет данных."])
        current_row += 2
    ws.cell(row=current_row, column=1, value="Pairwise Lag Analysis (пример для первой пары)")
    current_row += 1
    if len(tool.data.columns) >= 2:
        pair = list(combinations(tool.data.columns, 2))[0]
        col1, col2 = pair
        series1 = tool.data[col1].dropna().values
        series2 = tool.data[col2].dropna().values
        n = min(len(series1), len(series2))
        lag_metrics = {}
        for lag in range(1, 21):
            if n > lag:
                corr = np.corrcoef(series1[lag:], series2[:n-lag])[0, 1] if len(series1[lag:]) > 1 and len(series2[:n-lag]) > 1 else np.nan
                lag_metrics[lag] = corr
        if lag_metrics:
            lags = list(lag_metrics.keys())
            correlations = [lag_metrics[lag] for lag in lags]
            fig, ax = plt.subplots(figsize=(4, 3))
            ax.plot(lags, correlations, marker='o')
            ax.set_title(f"Lag Analysis: {col1}-{col2}")
            legend_text_pair = f"Пара: {col1}-{col2}, Метод: Lag Correlation"
            ax.text(0.5, 0.1, legend_text_pair, transform=ax.transAxes, fontsize=8, 
                    verticalalignment='bottom', bbox=dict(facecolor='white', alpha=0.5))
            buf_plag = BytesIO()
            plt.tight_layout()
            plt.savefig(buf_plag, format="png", dpi=100)
            buf_plag.seek(0)
            plt.close(fig)
            img_plag = Image(buf_plag)
            img_plag.width = 400
            img_plag.height = 300
            ws.add_image(img_plag, f"A{current_row}")
        else:
            ws.append(["Недостаточно данных для Pairwise Lag Analysis."])
        current_row += 20
        # (пример для первой пары)
        ws.cell(row=current_row, column=1, value="Extended Spectral Analysis (пример для первой пары)")
        current_row += 1
        title_es = f"Coherence {col1}-{col2}"
        buf_es = plot_coherence_vs_frequency(tool.data[col1], tool.data[col2], title_es, fs=100)
        img_es = Image(buf_es)
        img_es.width = 400
        img_es.height = 300
        ws.add_image(img_es, f"A{current_row}")
        current_row += 20
        # (пример для первой пары)
        ws.cell(row=current_row, column=1, value="Frequency Demonstration (пример для первой пары)")
        current_row += 1
        buf_fd = plot_coherence_vs_frequency(tool.data[col1], tool.data[col2], title_es, fs=100) 
        img_fd = Image(buf_fd)
        img_fd.width = 400
        img_fd.height = 300
        ws.add_image(img_fd, f"A{current_row}")
        current_row += 20
    else:
        ws.append(["Недостаточно столбцов для Pairwise Lag Analysis или Spectral Analysis."])
        current_row += 60 
    ws.cell(row=current_row, column=1, value="End of Combined Informational Analysis")
    logging.info("[Combined Informational Analysis] Лист сформирован.")

##############################################
#Combined Time Series (агрегированный + индивидуальные графики)
##############################################
def export_combined_ts_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined Time Series")
    ws.append(["Aggregated Time Series: Оригинальные и Нормализованные (на одном графике)"])
    buf_orig = tool.plot_time_series(tool.data, "Aggregated Original Time Series")
    img_orig = Image(buf_orig)
    img_orig.width = 600
    img_orig.height = 300
    ws.add_image(img_orig, "A2")
    
    ws.append([])
    buf_norm = tool.plot_time_series(tool.data_normalized, "Aggregated Normalized Time Series")
    img_norm = Image(buf_norm)
    img_norm.width = 600
    img_norm.height = 300
    ws.add_image(img_norm, "A10")
    
    ws.append(["Individual Time Series Plots:"])
    row = ws.max_row + 2
    for col in tool.data.columns:
        buf_ind = tool.plot_single_time_series(tool.data[col], f"Original: {col}")
        img_ind = Image(buf_ind)
        img_ind.width = 300
        img_ind.height = 200
        ws.add_image(img_ind, f"A{row}")
        buf_ind_norm = tool.plot_single_time_series(tool.data_normalized[col], f"Normalized: {col}")
        img_ind_norm = Image(buf_ind_norm)
        img_ind_norm.width = 300
        img_ind_norm.height = 200
        ws.add_image(img_ind_norm, f"E{row}")
        row += 15
    logging.info("[Combined Time Series] Лист сформирован.")

##############################################
# Combined FFT (агрегированный + индивидуальные графики)
##############################################
def export_all_fft_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined FFT")
    ws.append(["Combined FFT Analysis (Aggregated) - Original"])
    buf_fft_orig = tool.plot_fft(tool.data, "Aggregated Original FFT")
    img_fft_orig = Image(buf_fft_orig)
    img_fft_orig.width = 600
    img_fft_orig.height = 400
    ws.add_image(img_fft_orig, "A2")
    
    ws.append([])
    ws.append(["Combined FFT Analysis (Aggregated) - Normalized"])
    buf_fft_norm = tool.plot_fft(tool.data_normalized, "Aggregated Normalized FFT")
    img_fft_norm = Image(buf_fft_norm)
    img_fft_norm.width = 600
    img_fft_norm.height = 400
    ws.add_image(img_fft_norm, "A20")
    
    ws.append(["Individual FFT Analysis:"])
    row = ws.max_row + 2
    for col in tool.data.columns:
        buf_fft_ind = tool.plot_single_fft(tool.data[col], f"Original FFT: {col}")
        img_fft_ind = Image(buf_fft_ind)
        img_fft_ind.width = 300
        img_fft_ind.height = 200
        ws.add_image(img_fft_ind, f"A{row}")
        buf_fft_ind_norm = tool.plot_single_fft(tool.data_normalized[col], f"Normalized FFT: {col}")
        img_fft_ind_norm = Image(buf_fft_ind_norm)
        img_fft_ind_norm.width = 300
        img_fft_ind_norm.height = 200
        ws.add_image(img_fft_ind_norm, f"E{row}")
        row += 15
    logging.info("[Combined FFT] Лист сформирован.")

##############################################
# Создание оглавления с гиперссылками
##############################################
def create_table_of_contents(wb: Workbook):
    if "Table of Contents" in wb.sheetnames:
        old_sheet = wb["Table of Contents"]
        wb.remove(old_sheet)
    toc = wb.create_sheet("Table of Contents", 0)
    row = 1
    for sheet_name in wb.sheetnames:
        if sheet_name == "Table of Contents":
            continue
        link = f"#{sheet_name}!A1"
        cell = toc.cell(row=row, column=1)
        cell.value = sheet_name
        cell.hyperlink = link
        cell.style = "Hyperlink"
        row += 1

##############################################
# Класс BigMasterTool 
##############################################
class BigMasterTool:
    def __init__(self, data: pd.DataFrame = None, enable_experimental: bool = False) -> None:
        if data is not None:
            data = data.loc[:, (data != data.iloc[0]).any()]
            
            self.data = data.copy()
            # Гарантируем, что колонки называются c1, c2, ... и нет служебных
            numeric_cols = [c for c in self.data.columns if pd.api.types.is_numeric_dtype(self.data[c])] # ИСПРАВЛЕНИЕ ЗДЕСЬ
            self.data = self.data[numeric_cols]
            self.data.columns = [f'c{i+1}' for i in range(self.data.shape[1])]
        else:
            self.data = pd.DataFrame()

        self.data_normalized: pd.DataFrame = pd.DataFrame()
        self.results: dict = {}
        self.lag_results: dict = {}
        self.fft_results: dict = {}
        self.data_type: str = 'unknown'
        self.enable_experimental = enable_experimental
        self.lag_ranges = {v: range(1, 21) for v in method_mapping}
        self.undirected_methods = [
            "correlation_full",
            "correlation_partial",
            "h2_full",
            "h2_partial",
            "mutinf_full",
            "mutinf_partial",
            "coherence_full",
        ]
        self.directed_methods = [
            "correlation_directed",
            "h2_directed",
            "granger_full",
            "granger_partial",
            "granger_directed",
            "te_full",
            "te_partial",
            "te_directed",
            "ah_full",
            "ah_partial",
            "ah_directed",
        ]

    def load_data_excel(self, filepath: str, log_transform=False, remove_outliers=True, normalize=True, fill_missing=True, check_stationarity=False) -> pd.DataFrame:
        self.data = load_or_generate(filepath, log_transform, remove_outliers, normalize, fill_missing, check_stationarity)
        self.raw_data = self.data.copy()
        self.data_normalized = self.data.copy() # Инициализируем data_normalized СРАЗУ после загрузки
        if self.data.shape[0] < self.data.shape[1]:
            self.data = self.data.T
            self.data.columns = [f'c{i+1}' for i in range(self.data.shape[1])] 
        self.data = self.data.fillna(self.data.mean())
        self.data_type = 'file'
        logging.info(f"[BigMasterTool] Данные загружены, shape = {self.data.shape}.")
        return self.data

    def normalize_data(self):
        if self.data is None or self.data.empty or self.data.shape[1] == 0:
            logging.warning("[BigMasterTool] normalize_data: нет данных для нормализации.")
            self.data_normalized = pd.DataFrame()
            return

        cols_to_norm = [c for c in self.data.columns if pd.api.types.is_numeric_dtype(self.data[c])]
        if not cols_to_norm:
            self.data_normalized = self.data.copy() 
            logging.warning("[BigMasterTool] normalize_data: нет числовых колонок для нормализации.")
            return

        sc = StandardScaler()
        self.data_normalized = self.data.copy()
        self.data_normalized[cols_to_norm] = sc.fit_transform(self.data[cols_to_norm])
        logging.info("[BigMasterTool] Данные нормализованы.")


    def detect_outliers(self, col: str, thresh: float = 3, use_normalized: bool = False) -> np.ndarray:
        data = self.data_normalized if use_normalized else self.data
        if col not in data.columns or not pd.api.types.is_numeric_dtype(data[col]):
            return np.array([])
        arr = data[col].dropna().values
        if len(arr) == 0:
            return np.array([])
        z_scores = np.abs(stats.zscore(arr))
        return np.where(z_scores > thresh)[0]

    def remove_outliers(self, thresh: float = 5) -> None:
        if self.data.empty: return
        orig_shape = self.data.shape
        all_outl_indices = set()
        for col in self.data.columns:
            outl = self.detect_outliers(col, thresh=thresh, use_normalized=False)
            if len(outl) > 0:
                logging.info(f"[Outliers] '{col}': найдено {len(outl)} выбросов (z>{thresh}).")
                all_outl_indices.update(self.data.index[outl].tolist())
            else:
                logging.info(f"[Outliers] '{col}': выбросов нет.")
        
        if all_outl_indices:
            self.data = self.data.drop(list(all_outl_indices)).reset_index(drop=True)
            new_shape = self.data.shape
            logging.info(f"[Outliers] Размер данных {orig_shape} => {new_shape}.")
            self.normalize_data()
        else:
            logging.info("[Outliers] Данные без изменений.")

    def optimize_lag(self, variant: str, candidate_lags: range = range(1, 21)) -> dict:
        original_data = self.data_normalized.copy() 
        lag_metrics = {}
        for lag in tqdm(candidate_lags, desc=f"Optimizing lag for {variant}"):
            try:
                mat = compute_connectivity_variant(original_data.copy(), variant, lag)
                if mat is not None and hasattr(mat, "shape") and mat.shape == (original_data.shape[1], original_data.shape[1]):

                    if mat.shape[0] > 1:
                        metric = np.nanmean(np.abs(mat[~np.eye(mat.shape[0], dtype=bool)]))
                    else: 
                        metric = np.nan
                else:
                    metric = np.nan
                lag_metrics[lag] = (metric, mat)
            except Exception as ex:
                logging.error(f"[Lag] Ошибка {variant} lag={lag}: {ex}")
                num_cols = original_data.shape[1] if not original_data.empty else 0
                lag_metrics[lag] = (np.nan, np.full((num_cols, num_cols), np.nan))
        return lag_metrics 

    def analyze_lags(self, variant: str, candidate_lags: range = None) -> dict:
        c_lags = self.lag_ranges.get(variant, range(1, 21)) if candidate_lags is None else candidate_lags
        return self.optimize_lag(variant, c_lags)

    def compute_all_matrices(self):
        if self.data_normalized.empty: self.normalize_data()
        if self.data_normalized.empty:
            logging.warning("[Matrices] Нет данных для вычисления матриц.")
            return

        self.all_control_sets = [list(s) for s in powerset(self.data_normalized.columns)]
        self.connectivity_matrices = {}
        for method in method_mapping.keys():
            self.connectivity_matrices[method] = {}
            for S in self.all_control_sets:
                mat = compute_connectivity_variant(self.data_normalized.copy(), method, lag=1, control=S)
                if mat is not None:
                    self.connectivity_matrices[method][frozenset(S)] = mat
        logging.info("[Matrices] Все матрицы вычислены.")

    def prepare_pairs(self):
        if self.data.empty: return
        self.undirected_pairs = list(combinations(self.data.columns, 2))
        self.directed_pairs = list(permutations(self.data.columns, 2))
        self.undirected_rows = []
        for pair in self.undirected_pairs:
            others = [c for c in self.data.columns if c not in pair]
            for S in (list(s) for s in powerset(others)):
                self.undirected_rows.append((pair, S))
        self.directed_rows = []
        for pair in self.directed_pairs:
            others = [c for c in self.data.columns if c not in pair]
            for S in (list(s) for s in powerset(others)):
                self.directed_rows.append((pair, S))
        logging.info("[Pairs] Пары сформированы.")

    def get_undirected_value(self, mat, var1, var2, indices):
        if mat is None: return np.nan
        i, j = indices[var1], indices[var2]
        return mat[min(i, j), max(i, j)]

    def get_directed_value(self, mat, src, tgt, indices):
        if mat is None: return np.nan
        i, j = indices[src], indices[tgt]
        return mat[i, j]


    def run_all_methods(self) -> None:
        self.normalize_data()
        if self.data_normalized.empty:
            logging.warning("[RunAll] Нет данных для выполнения анализа.")
            return

        self.results = {}
        self.lag_results = {}
        if self.enable_experimental:
            self.fft_results = sliding_fft_analysis(
                self.data_normalized,
                window_size=min(200, len(self.data_normalized) // 2),
                overlap=min(100, len(self.data_normalized) // 4),
            )
        else:
            self.fft_results = {}
        
        # база, лаг 1
        for variant in method_mapping.keys():
            self.results[variant] = compute_connectivity_variant(self.data_normalized.copy(), variant, lag=1)
        
        # оптимиз лаг для директед
        directed_methods_for_lag_analysis = [
            "correlation_directed",
            "h2_directed",
            "granger_full",
            "granger_partial",
            "granger_directed",
            "te_full",
            "te_partial",
            "te_directed",
            "ah_full",
            "ah_partial",
            "ah_directed",
        ]
        for variant in directed_methods_for_lag_analysis:
            if variant in method_mapping: # d
                 self.lag_results[variant] = self.analyze_lags(variant, self.lag_ranges.get(variant, range(1, 21)))
            else:
                logging.warning(f"[RunAll] Метод {variant} не найден в method_mapping, пропуск анализа лагов.")

        self.compute_all_matrices()
        self.prepare_pairs()
        logging.info("[RunAll] Все методы завершены.")

    def analyze_sliding_windows(self, variant: str, window_size: int = 100, overlap: int = 50, threshold: float = 0.2) -> dict:
        if self.data_normalized.empty:
            return {}
        if not self.enable_experimental:
            return {}
        actual_window_size = min(window_size, len(self.data_normalized))
        actual_overlap = min(overlap, actual_window_size // 2) 
        return analyze_sliding_windows_with_metric(self.data_normalized, variant, actual_window_size, actual_overlap)

    def sliding_window_pairwise_analysis(self, method: str, window_size: int = 50, overlap: int = 25) -> dict:
        if self.data_normalized.empty or not self.enable_experimental:
            return {}
        actual_window_size = min(window_size, len(self.data_normalized))
        actual_overlap = min(overlap, actual_window_size // 2)
        return sliding_window_pairwise_analysis(self.data_normalized, method, actual_window_size, actual_overlap)

    def plot_lag_metrics(self, variant: str, lag_results: dict, legend_text: str = "") -> BytesIO:
        valid = [(l, lag_results[l][0]) for l in sorted(lag_results.keys()) if not np.isnan(lag_results[l][0])]
        if not valid:
            return BytesIO()
        lags, metrics = zip(*valid)
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.plot(lags, metrics, marker='o', linestyle='-', color='b', label=legend_text)
        ax.set_title(f"Lag Metrics for {variant.upper()} (используемые лаги: {min(lags)}-{max(lags)})")
        ax.set_xlabel("Lag")
        ax.set_ylabel("Metric")
        ax.grid(True)
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_sliding_window_comparison(self, sw_results: dict, legend_text: str = "") -> BytesIO:
        if not sw_results: return BytesIO()
        positions = sorted(sw_results.keys())
        metrics = [sw_results[pos][2] for pos in positions]
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(positions, metrics, marker='o', linestyle='-', color='m')
        ax.set_title("Sliding Window Metric vs. Window Start (окно = 50 точек)")
        ax.set_xlabel("Window Start Index")
        ax.set_ylabel("Metric")
        ax.grid(True)
        if legend_text:
            ax.text(0.05, 0.95, legend_text, transform=ax.transAxes, fontsize=8,
                    verticalalignment='top', bbox=dict(facecolor='white', alpha=0.5))
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_all_methods_lag_comparison(self, lag_results_dict: dict) -> BytesIO:
        fig, ax = plt.subplots(figsize=(10, 6))
        for variant, res in lag_results_dict.items():
            valid = [(l, res[l][0]) for l in sorted(res.keys()) if not np.isnan(res[l][0])]
            if valid:
                lags, metrics = zip(*valid)
                ax.plot(lags, metrics, marker='o', label=variant.upper())
        ax.set_title("Lag Analysis Comparison")
        ax.set_xlabel("Lag")
        ax.set_ylabel("Metric")
        ax.legend()
        ax.grid(True)
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def assess_significance(self,
                        method: str,
                        n_permutations: int = 100,
                        alpha: float = 0.05):
 
        orig = compute_connectivity_variant(self.data_normalized.copy(), method, lag=1)
        if orig is None or orig.shape[0] < 2:
            return None, None

        n = orig.shape[0]
        idx_i, idx_j = np.triu_indices(n, k=1)
        n_pairs = len(idx_i)
        
        perm_vals = np.zeros((n_permutations, n_pairs), dtype=float)

        for k in range(n_permutations):
            permuted = self.data_normalized.apply(np.random.permutation, axis=0)
            pm = compute_connectivity_variant(permuted.copy(), method, lag=1)
            if pm is None or pm.shape != orig.shape:
                perm_vals[k, :] = np.nan
            else:
                perm_vals[k, :] = np.abs(pm[idx_i, idx_j])

        orig_vals = np.abs(orig[idx_i, idx_j])
        
        # Здесь могут быть NaN.
        if np.any(np.isnan(orig_vals)) or n_permutations == 0:
            p_matrix = np.full_like(orig, np.nan)
            sig = np.full_like(orig, False, dtype=bool)
            return sig, p_matrix

        p_vals = (
            (perm_vals >= orig_vals[None, :]).sum(axis=0) + 1
        ) / (n_permutations + 1)

        sig = np.zeros_like(orig, dtype=bool)
        for (i,j), p in zip(zip(idx_i, idx_j), p_vals):
            sig[i, j] = sig[j, i] = (p < alpha)

        np.fill_diagonal(sig, False)

        p_matrix = np.zeros_like(orig, dtype=float)
        for (i,j), p in zip(zip(idx_i, idx_j), p_vals):
            p_matrix[i, j] = p_matrix[j, i] = p

        return sig, p_matrix


    def compute_hurst_rs(self, series) -> float:
        try:
            H, _, _ = compute_Hc(series.dropna().values, kind='change', simplified=True)
            return H
        except Exception as ex:
            logging.error(f"[Hurst RS] Ошибка: {ex}")
            return np.nan

    def compute_hurst_dfa(self, series) -> float:
        try:
            H = nolds.dfa(series.dropna().values)
            return H
        except Exception as ex:
            logging.error(f"[Hurst DFA] Ошибка: {ex}")
            return np.nan

    def compute_hurst_aggregated_variance(self, series, max_n=100) -> float:
        try:
            arr = np.array(series.dropna())
            N = len(arr)
            if N < max_n:
                return np.nan
            m_vals = np.arange(1, min(max_n+1, N//2))
            variances = []
            for m in m_vals:
                nb = N // m
                if nb > 0:
                    reshaped = arr[:nb*m].reshape(nb, m)
                    block_means = reshaped.mean(axis=1)
                    if len(block_means) > 1:
                        variances.append(np.var(block_means))
            if not variances: return np.nan
            log_m = np.log10(m_vals[:len(variances)])
            log_var = np.log10(variances)
            slope, _ = np.polyfit(log_m, log_var, 1)
            H = 1 - slope/2
            return H
        except Exception as ex:
            logging.error(f"[Hurst AggVar] Ошибка: {ex}")
            return np.nan

    def compute_hurst_wavelet(self, series) -> float:
        try:
            arr = np.array(series.dropna())
            N = len(arr)
            if N < 50:
                return np.nan
            yf_arr = fft(arr)
            freqs = np.fft.fftfreq(N)
            psd = np.abs(yf_arr)**2
            idx = freqs > 0
            freqs = freqs[idx]
            psd = psd[idx]
            if len(freqs) < 2: return np.nan
            log_freqs = np.log10(freqs)
            log_psd = np.log10(psd)
            slope, _ = np.polyfit(log_freqs, log_psd, 1)
            H = (1 - slope)/2
            return H
        except Exception as ex:
            logging.error(f"[Hurst Wavelet] Ошибка: {ex}")
            return np.nan
    #графики?
    def plot_autocorrelation(self, series: pd.Series, title: str, suppress_noise: bool = False, noise_threshold: float = 0.9, legend_text: str = "") -> BytesIO:
        fig, ax = plt.subplots(figsize=(6, 4))
        plot_acf(series.dropna(), ax=ax, lags=min(50, len(series)-1), zero=False, alpha=0.05, fft=True)
        
        ax.set_title(title + (" (с подавлением шума)" if suppress_noise else ""))
        if legend_text:
            ax.text(0.05, 0.95, legend_text, transform=ax.transAxes, fontsize=8,
                    verticalalignment='top', bbox=dict(facecolor='white', alpha=0.5))
        buf  = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_power_spectrum(self, series: pd.Series, title: str) -> BytesIO:
        N = len(series)
        if N < 2: return BytesIO()
        yf_arr = fft(series)
        freqs = np.fft.fftfreq(N)
        amp = np.abs(yf_arr)**2
        idx = freqs > 0
        freqs = freqs[idx]
        amp = amp[idx]
        if len(freqs) < 2: return BytesIO()
        log_freqs = np.log10(freqs)
        log_amp = np.log10(amp)
        slope, _ = np.polyfit(log_freqs, log_amp, 1)
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.loglog(freqs, amp, label=f"Slope={slope:.2f}")
        ax.set_title(title)
        ax.set_xlabel("Частота (log)")
        ax.set_ylabel("Мощность (log)")
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)
        return buf

    def detect_seasonality(self, series: pd.Series, threshold=0.2):
        N = len(series)
        if N < 2: return None, None
        yf_arr = fft(series)
        freqs = np.fft.fftfreq(N)
        amp = np.abs(yf_arr)**2
        idx = freqs > 0
        freqs = freqs[idx]
        amp = amp[idx]
        if amp.size == 0: return None, None
        peaks, props = find_peaks(amp, height=threshold*np.max(amp))
        if len(peaks) > 0:
            peak_freqs = freqs[peaks]
            periods = 1 / peak_freqs
            return peak_freqs, periods
        return None, None

    def export_hurst_sheet(self, wb: Workbook, method_name: str, compute_func, plot_func):
        ws = wb.create_sheet(method_name)
        ws.append(["Столбец", "Hurst Exponent"])
        row_num = 2
        for col in self.data.columns:
            s = self.data[col].dropna()
            if len(s) < 50:
                ws.append([col, "N/A (недостаточно данных)"])
                row_num += 1
                continue
            H = compute_func(s)
            if H is not None and not np.isnan(H):
                ws.append([col, f"{H:.3f}"])
                buf = plot_func(s, f"{method_name} - {col}")
                img = Image(buf)
                img.width = 400
                img.height = 300
                ws.add_image(img, f"D{row_num}")
            else:
                ws.append([col, "N/A"])
            row_num += 1

    def export_seasonality_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Seasonality")
        ws.append(["Столбец", "Пиковые частоты", "Периоды", "Пояснение"])
        for col in self.data.columns:
            s = self.data[col].dropna()
            freq, periods = self.detect_seasonality(s)
            if freq is not None and np.asarray(freq).size > 0 and periods is not None and np.asarray(periods).size > 0:
                freq_str = ", ".join([f"{f:.3f}" for f in freq])
                period_str = ", ".join([f"{p:.1f}" for p in periods])
                explanation = f"Сезонность обнаружена, период ≈ {np.median(periods):.1f}"
                ws.append([col, freq_str, period_str, explanation]) 
            else:
                ws.append([col, "Нет пиков", "Нет пиков", "Сезонность не обнаружена"])
                
    def export_undirected_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Undirected Methods")
        headers = ["Pair", "Control Set"] + self.undirected_methods
        ws.append(headers)
        indices = {c: i for i, c in enumerate(self.data.columns)}
        fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        fill_pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        for row_idx, row in enumerate(self.undirected_rows, start=2):
            pair, S = row
            pair_str = "-".join(pair)
            S_str = ",".join(S) if S else "None"
            vals = [pair_str, S_str]
            for method in self.undirected_methods:
                mat = self.connectivity_matrices[method].get(frozenset(S), None)
                if mat is not None:
                    v = self.get_undirected_value(mat, pair[0], pair[1], indices)
                    vals.append(fmt_val(v))
                else:
                    vals.append("N/A")
            ws.append(vals)
            for col_idx, method in enumerate(self.undirected_methods, start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value != "N/A":
                    try:
                        v = float(cell.value)
                        if "granger" in method.lower():
                            cell.fill = fill_green if abs(v) < 0.05 else fill_pink
                        else:
                            cell.fill = fill_green if abs(v) > 0.2 else fill_pink
                    except:
                        pass
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

    def export_directed_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Directed Methods")
        headers = ["Directed Pair", "Control Set"] + self.directed_methods
        ws.append(headers)
        indices = {c: i for i, c in enumerate(self.data.columns)}
        fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        fill_pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        for row_idx, row in enumerate(self.directed_rows, start=2):
            pair, S = row
            pair_str = f"{pair[0]}->{pair[1]}"
            S_str = ",".join(S) if S else "None"
            vals = [pair_str, S_str]
            for method in self.directed_methods:
                mat = self.connectivity_matrices[method].get(frozenset(S), None)
                if mat is not None:
                    v = self.get_directed_value(mat, pair[0], pair[1], indices)
                    vals.append(fmt_val(v))
                else:
                    vals.append("N/A")
            ws.append(vals)
            for col_idx, method in enumerate(self.directed_methods, start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value != "N/A":
                    try:
                        v = float(cell.value)
                        if "granger" in method.lower():
                            cell.fill = fill_green if abs(v) < 0.05 else fill_pink
                        else:
                            cell.fill = fill_green if abs(v) > 0.2 else fill_pink
                    except:
                        pass
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

    def select_lag_metrics(self, lag_res: dict, variant: str):
        valid = [l for l in lag_res if not np.isnan(lag_res[l][0])]
        if not valid:
            return None, None, None
        metrics = [lag_res[l][0] for l in valid]
        is_granger = "granger" in variant.lower()
        if is_granger: # чем меньше р, тем ЛУЧШЕ Для грйнджера. грейнджер хочет МАЛЕНЬКОЕ р, НЕЕ МЕНЯТЬ ТУТ
            best = min(valid, key=lambda l: lag_res[l][0])
            worst = max(valid, key=lambda l: lag_res[l][0])
        else: 
            best = max(valid, key=lambda l: lag_res[l][0])
            worst = min(valid, key=lambda l: lag_res[l][0])
        median = min(valid, key=lambda l: abs(lag_res[l][0] - np.median(metrics)))
        return best, median, worst

    def export_method_sheet(self, wb: Workbook, variant: str, threshold: float, window_size: int, overlap: int) -> None:
        directed_flag = ("partial" in variant or "directed" in variant)
        if "granger" in variant.lower():
            granger_threshold = 0.05
            invert_threshold = True
        else:
            granger_threshold = threshold
            invert_threshold = False
        ws = wb.create_sheet(variant.upper() + " Results")
        ws.append([f"Метод: {variant.upper()} (Лаг = 1, окно = {window_size})"])
        full_mat = compute_connectivity_variant(self.data_normalized, variant, lag=1)
        if full_mat is None or not hasattr(full_mat, "shape") or (full_mat.shape[0] != self.data_normalized.shape[1] or full_mat.shape[1] != self.data_normalized.shape[1]):
             ws.append([f"{variant.upper()}: Метод не работает для этих данных или вернул матрицу некорректного размера."])
             return
        ws.append(["Полная матрица:"])
        legend_text = f"Лаг=1, Окно={window_size}"
        add_method_to_sheet(ws, ws.max_row + 1, f"{variant.upper()} Full Matrix", full_mat, directed=directed_flag, legend_text=legend_text)
        buf_heat_full = plot_heatmap(full_mat, f"{variant.upper()} Heatmap (Full)", legend_text=legend_text)
        img_heat_full = Image(buf_heat_full)
        img_heat_full.width = 400
        img_heat_full.height = 300
        ws.add_image(img_heat_full, "G2")
        buf_conn_full = plot_connectome(full_mat, f"{variant.upper()} Connectome (Full)",
                                        threshold=granger_threshold, directed=directed_flag, invert_threshold=invert_threshold, legend_text=legend_text)
        img_conn_full = Image(buf_conn_full)
        img_conn_full.width = 400
        img_conn_full.height = 400
        ws.add_image(img_conn_full, "M2")
        ws.append(["--- Медианный и Лучший лаг ---"])
        lag_res = self.lag_results.get(variant, {}) # исп предрассчит
        best_lag, median_lag, worst_lag = self.select_lag_metrics(lag_res, variant)
        if median_lag is not None and lag_res.get(median_lag) and lag_res[median_lag][1] is not None:
            med_mat = lag_res[median_lag][1]
            ws.append([f"Медианный лаг: {median_lag}"])
            legend_text_med = f"Лаг={median_lag}, Окно={window_size}"
            buf_med_heat = plot_heatmap(med_mat, f"{variant.upper()} Heatmap (Median Lag)", legend_text=legend_text_med)
            img_med_heat = Image(buf_med_heat)
            img_med_heat.width = 400
            img_med_heat.height = 300
            ws.add_image(img_med_heat, "A20")
            buf_med_conn = plot_connectome(med_mat, f"{variant.upper()} Connectome (Median Lag)", threshold, directed=directed_flag, legend_text=legend_text_med)
            img_med_conn = Image(buf_med_conn)
            img_med_conn.width = 400
            img_med_conn.height = 400
            ws.add_image(img_med_conn, "G20")
        else:
            ws.append([f"Медианный лаг: N/A"])

        if best_lag is not None and lag_res.get(best_lag) and lag_res[best_lag][1] is not None:
            best_mat = lag_res[best_lag][1]
            ws.append([f"Лучший лаг: {best_lag}"])
            legend_text_best = f"Лаг={best_lag}, Окно={window_size}"
            buf_best_heat = plot_heatmap(best_mat, f"{variant.upper()} Heatmap (Best Lag)", legend_text=legend_text_best)
            img_best_heat = Image(buf_best_heat)
            img_best_heat.width = 400
            img_best_heat.height = 300
            ws.add_image(img_best_heat, "M20")
            buf_best_conn = plot_connectome(best_mat, f"{variant.upper()} Connectome (Best Lag)", threshold, directed=directed_flag, legend_text=legend_text_best)
            img_best_conn = Image(buf_best_conn)
            img_best_conn.width = 400
            img_best_conn.height = 400
            ws.add_image(img_best_conn, "Q20")
        else:
            ws.append([f"Лучший лаг: N/A"])
            
        ws.append(["--- Анализ скользящих окон ---"])
        sw_summary = {}
        for w_size in [50, 100, 500]:
                # не брать слишком большое окно
            current_w_size = min(w_size, len(self.data_normalized) // 2) # надо хотя бы два
            if current_w_size < 10: # минимально разумный
                ws.append([f"Размер окна {w_size}: Недостаточно данных для скользящих окон."])
                continue

            sw_res = analyze_sliding_windows_with_metric(self.data_normalized, variant, window_size=current_w_size, overlap=current_w_size//2)
            legend_text_sw = f"Метод={variant}, Окно={current_w_size}"
            buf_sw = self.plot_sliding_window_comparison(sw_res, legend_text=legend_text_sw)
            img_sw = Image(buf_sw)
            img_sw.width = 400
            img_sw.height = 300
            cell = f"A{ws.max_row + 2}"
            ws.add_image(img_sw, cell)
            if sw_res:
                best_start = max(sw_res.keys(), key=lambda s: sw_res[s][2])
                sw_summary[current_w_size] = (best_start, sw_res[best_start])
                ws.append([f"Размер окна {current_w_size}: лучшее окно = {best_start}, метрика = {sw_res[best_start][2]:.3f}"])
            else:
                 ws.append([f"Размер окна {current_w_size}: Нет результатов для скользящих окон."])

        if sw_summary:
            best_overall = max(sw_summary.items(), key=lambda item: item[1][1][2])
            best_w_size = best_overall[0]
            best_w_start, best_w_data = best_overall[1]
            ws.append(["--- Лучшее окно среди всех ---"])
            ws.append([f"Размер окна: {best_w_size}, старт: {best_w_start}"])
            best_win_mat = best_w_data[1]
            ws.append(["Матрица лучшего окна:"])
            add_method_to_sheet(ws, ws.max_row + 1, f"{variant.upper()} Best Window Matrix", best_win_mat, directed=directed_flag, legend_text=f"Лаг=1, Окно={best_w_size}")
            buf_best_heat = plot_heatmap(best_win_mat, f"{variant.upper()} Heatmap (Best Window)", legend_text=f"Лаг=1, Окно={best_w_size}")
            img_best_heat = Image(buf_best_heat)
            img_best_heat.width = 400
            img_best_heat.height = 300
            ws.add_image(img_best_heat, "A40")
            buf_best_conn = plot_connectome(best_win_mat, f"{variant.upper()} Connectome (Best Window)", threshold, directed=directed_flag, legend_text=f"Лаг=1, Окно={best_w_size}")
            img_best_conn = Image(buf_best_conn)
            img_best_conn.width = 400
            img_best_conn.height = 400
            ws.add_image(img_best_conn, "G40")
        ws.append(["--- Конец листа ---"])

    def export_summary_sheet(self, wb: Workbook) -> None:
        ws_summary = wb.create_sheet("Summary")
        full_methods = [m for m in self.results if "partial" not in m and m != "granger_full"]
        # 
        if "granger_directed" in self.results and "granger_directed" not in full_methods:
            full_methods.append("granger_directed")        
        undirected_pairs = list(combinations(self.data.columns, 2))
        directed_pairs = list(permutations(self.data.columns, 2))
        
        headers = ["Название метода"]
        for p in undirected_pairs:
            p_str = f"{p[0]}-{p[1]}"
            headers.extend([f"{p_str} (мед)", f"{p_str} (луч)"])
        for p in directed_pairs:
            p_str = f"{p[0]}->{p[1]}"
            headers.extend([f"{p_str} (мед)", f"{p_str} (луч)"])
        headers.extend(["Лучший лаг", "Медианный лаг", "H_RS", "H_DFA", "H_AggVar", "H_Wavelet", "Seasonality Periods"])
        ws_summary.append(headers)
        
        fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        fill_pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        indices = {c: i for i, c in enumerate(self.data.columns)}
        
        for variant in full_methods:
            is_granger = "granger" in variant.lower()
            lag_res = self.lag_results.get(variant, {}) # берём превыч
            
            best_lag, median_lag, _ = self.select_lag_metrics(lag_res, variant)
            
            mat_med = lag_res.get(median_lag, [None, None])[1] if median_lag is not None else None
            mat_best = lag_res.get(best_lag, [None, None])[1] if best_lag is not None else None
            
            row = [variant]
            
            for pair in undirected_pairs:
                var1, var2 = pair
                if mat_med is not None and mat_best is not None:
                    i, j = indices.get(var1, -1), indices.get(var2, -1) 
                    if i != -1 and j != -1:
                        val_med = mat_med[min(i, j), max(i, j)] if mat_med.shape[0] > max(i,j) and mat_med.shape[1] > max(i,j) else np.nan
                        val_best = mat_best[min(i, j), max(i, j)] if mat_best.shape[0] > max(i,j) and mat_best.shape[1] > max(i,j) else np.nan
                    else:
                        val_med, val_best = np.nan, np.nan
                else:
                    val_med = np.nan
                    val_best = np.nan
                row.extend([fmt_val(val_med), fmt_val(val_best)])

            for pair in directed_pairs:
                src, tgt = pair
                if mat_med is not None and mat_best is not None:
                    i, j = indices.get(src, -1), indices.get(tgt, -1)
                    if i != -1 and j != -1:
                        val_med = mat_med[i, j] if mat_med.shape[0] > i and mat_med.shape[1] > j else np.nan
                        val_best = mat_best[i, j] if mat_best.shape[0] > i and mat_best.shape[1] > j else np.nan
                    else:
                        val_med, val_best = np.nan, np.nan
                else:
                    val_med = np.nan
                    val_best = np.nan
                row.extend([fmt_val(val_med), fmt_val(val_best)])
            
            row.extend([best_lag if best_lag is not None else 'N/A', median_lag if median_lag is not None else 'N/A'])
            
            hurst_vals_for_avg = []
            season_vals_for_avg = []
            for col in self.data.columns:
                s = self.data[col].dropna()
                H_RS = self.compute_hurst_rs(s)
                H_DFA = self.compute_hurst_dfa(s)
                H_AggVar = self.compute_hurst_aggregated_variance(s)
                H_Wavelet = self.compute_hurst_wavelet(s)
                
                hurst_vals_for_avg.extend([H_RS, H_DFA, H_AggVar, H_Wavelet])
                
                freq, periods = self.detect_seasonality(s)
                season_text = f"Есть сезонность, период ≈ {np.median(periods):.1f}" if freq is not None and periods is not None and len(periods) > 0 else "Нет сезонности"
                season_vals_for_avg.append(season_text)

            avg_hurst_values = [h for h in hurst_vals_for_avg if h is not None and not np.isnan(h)]
            avg_hurst_results = [fmt_val(np.mean(avg_hurst_values)) if avg_hurst_values else "N/A"] * 4
            
            avg_season_text = "; ".join(sorted(list(set(season_vals_for_avg)))) 
            
            row.extend(avg_hurst_results)
            row.append(avg_season_text)
            ws_summary.append(row)
            
            cur_row = ws_summary.max_row
            num_pair_cols = (len(undirected_pairs) + len(directed_pairs)) * 2
            for idx in range(2, 2 + num_pair_cols):
                cell = ws_summary.cell(row=cur_row, column=idx)
                if cell.value != "N/A":
                    try:
                        v = float(cell.value)
                        if is_granger:
                            cell.fill = fill_green if abs(v) < 0.05 else fill_pink
                        else:
                            cell.fill = fill_green if abs(v) > 0.2 else fill_pink
                    except:
                        pass
        
        for i, _ in enumerate(headers, start=1):
            max_len = 0
            for r in range(1, ws_summary.max_row + 1):
                cell_value = ws_summary.cell(row=r, column=i).value
                if cell_value is not None:
                    max_len = max(max_len, len(str(cell_value)))
            ws_summary.column_dimensions[get_column_letter(i)].width = max_len + 2

    def export_coefficients_sheet(self, wb: Workbook):
        export_coefficients_sheet(self, wb)

    def export_frequency_summary_sheet(self, wb: Workbook):
        export_frequency_summary_sheet(self, wb)

    def export_frequency_dependence_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Frequency Dependence")
        ws.append(["Пара", "Макс. когерентность", "Частота (Hz)"])
        for pair in combinations(self.data.columns, 2):
            col1, col2 = pair
            s1 = self.data[col1].dropna()
            s2 = self.data[col2].dropna()
            n = min(len(s1), len(s2))
            if n == 0:
                ws.append([f"{col1}-{col2}", "N/A", "N/A"])
                continue
            s1 = s1.values[:n]
            s2 = s2.values[:n]
            fs = 100
            freqs, cxy = coherence(s1, s2, fs=fs)
            if cxy.size == 0:
                ws.append([f"{col1}-{col2}", "N/A", "N/A"])
                continue
            max_idx = np.argmax(cxy)
            max_coh = cxy[max_idx]
            max_freq = freqs[max_idx]
            ws.append([f"{col1}-{col2}", f"{max_coh:.3f}", f"{max_freq:.3f}"])
            buf = plot_coherence_vs_frequency(self.data[col1], self.data[col2], f"Coherence {col1}-{col2}", fs=fs)
            img = Image(buf)
            img.width = 400
            img.height = 300
            cell = f"D{ws.max_row}"
            ws.add_image(img, cell)

    def export_big_excel(self, save_path: str = "AllMethods_Full.xlsx", threshold: float = 0.2, window_size: int = 100, overlap: int = 50,
                           log_transform=False, remove_outliers=True, normalize=True, fill_missing=True, check_stationarity=False) -> str:
        wb = Workbook()
        wb.remove(wb.active)
        add_raw_data_sheet(wb, self.data) 
        self.export_summary_sheet(wb)     
        self.export_undirected_sheet(wb)     
        self.export_directed_sheet(wb)         
        ws_diag = wb.create_sheet("Data & Diagnostics")
        ws_diag.append(["Диагностика (Original) с информацией о сезонности"])
        for c in self.data.columns:
            adf_stat, adf_pv = self.test_stationarity(c, use_normalized=False)
            outl = self.detect_outliers(c, thresh=3, use_normalized=False)
            freq, periods = self.detect_seasonality(self.data[c].dropna())
            season_text = f"Период ≈ {np.median(periods):.1f}" if freq is not None and periods is not None and len(periods) > 0 else "Нет сезонности"
            ws_diag.append([f"Столбец: {c}"])
            if adf_stat is not None and adf_pv is not None:
                ws_diag.append([f"ADF: {adf_stat:.6f} (p={adf_pv:.6f})"])
            else:
                ws_diag.append(["ADF: N/A"])
            ws_diag.append([f"Выбросов: {len(outl)}, Пример: {list(outl)[:10]}"])
            ws_diag.append([f"Сезонность: {season_text}"])
            buf_ps = self.plot_power_spectrum(self.data[c].dropna(), f"Power Spectrum {c}")
            img_ps = Image(buf_ps)
            img_ps.width = 400
            img_ps.height = 300
            cell = f"D{ws_diag.max_row - 1}"
            ws_diag.add_image(img_ps, cell)
            ws_diag.append([])
        
        # Combined Time Series – агрегированный и индивидуальные графики
        export_combined_ts_sheet(self, wb)
        
        # Combined FFT – агрегированный и индивидуальные графики
        export_all_fft_sheet(self, wb)
        
        # Объединённый информационный лист
        export_combined_informational_sheet(self, wb)
        
        # --- ЛИСТЫ методов отдельно ---
        for variant in method_mapping.keys():
            self.export_method_sheet(wb, variant, threshold, window_size, overlap)
        
        self.export_hurst_sheet(wb, "Hurst_RS", self.compute_hurst_rs, self.plot_autocorrelation)
        self.export_hurst_sheet(wb, "Hurst_DFA", self.compute_hurst_dfa, self.plot_power_spectrum)
        self.export_hurst_sheet(wb, "Hurst_AggVar", self.compute_hurst_aggregated_variance, self.plot_autocorrelation)
        self.export_hurst_sheet(wb, "Hurst_Wavelet", self.compute_hurst_wavelet, self.plot_power_spectrum)
        self.export_seasonality_sheet(wb)
        self.export_coefficients_sheet(wb)
        self.export_frequency_summary_sheet(wb)
        self.export_frequency_dependence_sheet(wb)
        export_individual_ac_ph_sheet(self, wb)
        export_entropy_sheet(self, wb)
        
        if "Combined Time Series1" in wb.sheetnames:
            ws_to_remove = wb["Combined Time Series1"]
            wb.remove(ws_to_remove)
        
        create_table_of_contents(wb)
        
        wb.save(save_path)
        logging.info(f"[Export] Excel файл сохранён: {save_path}")
        return save_path

    def test_stationarity(self, col: str, use_normalized: bool = False):
        data = self.data_normalized if use_normalized else self.data
        if col not in data.columns or not pd.api.types.is_numeric_dtype(data[col]):
            return None, None
        arr = data[col].dropna()
        if len(arr) < 2:
            return None, None
        try:
            adf_res = adfuller(arr)
            logging.debug(f"[Stationarity] {col}: ADF = {adf_res[0]:.6f}, p = {adf_res[1]:.6f}")
            return adf_res[0], adf_res[1]
        except Exception as ex:
            logging.error(f"Ошибка ADF для {col}: {ex}")
            return None, None

    def evaluate_noise(self, col: str, use_normalized: bool = False):
        return 0, 0  # ЗАГЛУШКА

    def plot_time_series(self, data: pd.DataFrame, title: str) -> BytesIO:
        fig, ax = plt.subplots(figsize=(8, 4))
        for c in data.columns:
            if pd.api.types.is_numeric_dtype(data[c]):
                ax.plot(data[c].dropna(), label=c)
        ax.set_title(title)
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_single_time_series(self, series: pd.Series, title: str) -> BytesIO:
        fig, ax = plt.subplots(figsize=(4, 3))
        if pd.api.types.is_numeric_dtype(series):
            ax.plot(series.dropna(), label=series.name)
        ax.set_title(title)
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_fft(self, data: pd.DataFrame, title: str) -> BytesIO:
        fig, ax = plt.subplots(figsize=(8, 4))
        for c in data.columns:
            if pd.api.types.is_numeric_dtype(data[c]):
                freqs, amplitude, _, _ = plt_fft_analysis(data[c])
                if freqs.size > 0:
                    ax.plot(freqs, amplitude, label=c)
        ax.set_title(title)
        ax.set_xlabel("Частота")
        ax.set_ylabel("Амплитуда")
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_single_fft(self, series: pd.Series, title: str) -> BytesIO:
        freqs, amplitude, phase, peaks = plt_fft_analysis(series)
        fig, ax = plt.subplots(figsize=(4, 3))
        if freqs.size > 0:
            ax.plot(freqs, amplitude, label=series.name)
            ax.plot(freqs[peaks], amplitude[peaks], "x", label="Peaks")
        ax.set_title(title)
        ax.set_xlabel("Частота")
        ax.set_ylabel("Амплитуда")
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)
        return buf

    def get_heatmaps_for_summary(self, methods_to_get: dict, annotate: bool = False, lag_params: dict = None) -> dict:
        """ Модифицированная версия с опцией аннотации и кастомными лагами. """
        if self.data_normalized.empty: self.normalize_data()
        if lag_params is None:
            lag_params = {}

        generated_heatmaps = {}
        df_for_analysis = self.data_normalized[[c for c in self.data_normalized.columns if c.startswith('c')]]
        if df_for_analysis.empty: return {k: None for k in methods_to_get}

        for report_label, method_variant in methods_to_get.items():
            # ПРЕДЗАДАННЫЕ КОНТРОЛЬНЫЕ для этих там
            control_vars = ['c3'] if 'partial' in method_variant else None
            
            lag = lag_params.get(method_variant, DEFAULT_MAX_LAG)
            
            matrix = compute_connectivity_variant(df_for_analysis, method_variant, lag=lag, control=control_vars)
            
            title = f"{report_label}"
            legend_text = f"Lag={lag}"
            image_buffer = plot_heatmap(matrix, title, legend_text=legend_text, annotate=annotate)
            generated_heatmaps[report_label] = image_buffer
        return generated_heatmaps
    def get_connectomes_for_summary(self, methods_to_get: dict, lag_params: dict = None) -> dict:
      
        if self.data_normalized.empty: self.normalize_data()
        if lag_params is None:
            lag_params = {}

        generated_connectomes = {}
        df_for_analysis = self.data_normalized[[c for c in self.data_normalized.columns if c.startswith('c')]]
        if df_for_analysis.empty: return {k: None for k in methods_to_get}

        for report_label, method_variant in methods_to_get.items():
            control_vars = ['c3'] if 'partial' in method_variant else None
            
            lag = lag_params.get(method_variant, DEFAULT_MAX_LAG)
            
            matrix = compute_connectivity_variant(df_for_analysis, method_variant, lag=lag, control=control_vars)

            if matrix is not None:
                is_directed = "directed" in method_variant or "partial" in method_variant
                is_granger = "granger" in method_variant
                threshold = 0.05 if is_granger else 0.5 
                invert_threshold = True if is_granger else False

                title = f"{report_label}"
                legend_text = f"Lag={lag}"
                image_buffer = plot_connectome(matrix, title, threshold=threshold, directed=is_directed, invert_threshold=invert_threshold, legend_text=legend_text)
                generated_connectomes[report_label] = image_buffer
            else:
                generated_connectomes[report_label] = None
        return generated_connectomes


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compute connectivity measures for multivariate time series."
    )
    parser.add_argument(
        "input_file",
        help="Path to input CSV or Excel file with time series data",
    )
    parser.add_argument(
        "--lags",
        type=int,
        default=DEFAULT_MAX_LAG,
        help="Lag or model order (for Granger, TE, etc.)",
    )
    parser.add_argument(
        "--log",
        action="store_true",
        help="Apply logarithm transform to data (for positive-valued data)",
    )
    parser.add_argument(
        "--no-outliers",
        action="store_true",
        help="Disable outlier removal",
    )
    parser.add_argument(
        "--no-normalize",
        action="store_true",
        help="Disable normalization of data",
    )
    parser.add_argument(
        "--no-stationarity-check",
        action="store_true",
        help="Disable stationarity check (ADF test)",
    )
    parser.add_argument(
        "--graph-threshold",
        type=float,
        default=0.5,
        help="Threshold for graph edges (weight >= threshold)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel file path (defaults to TimeSeriesAnalysis/AllMethods_Full.xlsx)",
    )
    parser.add_argument(
        "--quiet-warnings",
        action="store_true",
        help="Suppress most warnings for cleaner CLI output.",
    )
    parser.add_argument(
        "--experimental",
        action="store_true",
        help="Enable experimental sliding-window analyses.",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    configure_warnings(quiet=args.quiet_warnings)

    filepath = os.path.abspath(args.input_file)
    output_path = args.output or os.path.join(save_folder, "AllMethods_Full.xlsx")
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    tool = BigMasterTool(enable_experimental=args.experimental)
    tool.lag_ranges = {v: range(1, args.lags + 1) for v in method_mapping}
    tool.load_data_excel(
        filepath,
        log_transform=args.log,
        remove_outliers=not args.no_outliers,
        normalize=not args.no_normalize,
        fill_missing=True,
        check_stationarity=not args.no_stationarity_check,
    )
    tool.run_all_methods()
    tool.export_big_excel(
        output_path,
        threshold=args.graph_threshold,
        window_size=100,
        overlap=50,
        log_transform=args.log,
        remove_outliers=not args.no_outliers,
        normalize=not args.no_normalize,
        fill_missing=True,
        check_stationarity=not args.no_stationarity_check,
    )

    print("Анализ завершён, результаты сохранены в:", output_path)
